from openpyxl import load_workbook
from openpyxl import Workbook
import os

#需要设置分名字系统
#单独设置分值


def sayhello():
    print("程序开始执行")
    filename = input("请输入文件路径: ")
    address = filename
    #print("检测到文件路径为: "+address)
    return address


def askinfo_start():
    info_start = []
    row_start = input("请输入起始行： ")
    if row_start == None:
        row_start = 1
    info_start.append(row_start)
    column_start = input("请输入起始列：")
    if column_start == None:
        column_start = 1
    info_start.append(column_start)
    return info_start


def getanswer():
    answeraddress = input("请输入答案路径： ")
    #print("检测到答案地址为： "+answeraddress)
    wb = load_workbook(answeraddress)
    ws = wb.get_sheet_by_name("Sheet1")
    answer = []
    column = 1
    checkpoint = True
    while checkpoint:
        a = ws.cell(column, 1).value
        if(a == None):
            break
        else:
            answer.append(a)
        column = column+1
    print("答案是：")
    print(answer)
    return answer


def outputexcel(grade=[[], [], []]):  # 建立新的表格
    row = 1
    newwb = Workbook()
    newws = newwb.active
    for name in grade[0]:
        newws.cell(row, 1).value = name
        row += 1
    row = 1
    for grade_final in grade[1]:
        newws.cell(row, 2).value = grade_final
        row += 1
    row = 1
    for answer in grade[2]:
        column = 3
        for answer_s in answer:
            newws.cell(row, column).value = answer_s
            column += 1
        row += 1
    address = input("请输入输出地址：")
    newwb.save(address)


def dealwithname(grade):
    name_num = 0
    for name in grade[0]:
        name.rstrip()
        index = ["的", "妈", "爸", "家长","老师","爷爷","外婆","姐姐","奶奶","外公"]
        for a in index:
            if name.find(a) >= 0:
                newname = name.split(a)
                grade[0][name_num] = newname[0]
                break
        name_num += 1
    print(grade)
    return grade


def checkduplication(m_grade):  # 处理重复的数据
    pointer = 0
    name_check_time = 0
    student_del = []
    while name_check_time < len(m_grade[0]):
        student = []
        student.append(m_grade[0][pointer])
        student.append(pointer)
        grade_check_time = 0
        check_pointer = pointer+1
        a = len(m_grade[0])
        while grade_check_time < (len(m_grade[0])-pointer-1):
            if student[0] == m_grade[0][check_pointer]:
                if student[1] < m_grade[1][check_pointer]:
                    m_grade[1][pointer] = m_grade[1][check_pointer]
                    m_grade[2][pointer] = m_grade[2][check_pointer]
                    student_del.append(check_pointer)
                else:
                    student_del.append(check_pointer)
            grade_check_time += 1
            check_pointer += 1
        name_check_time += 1
        pointer += 1
        time = 0
        for delete in student_del:
            del m_grade[0][delete-time]
            del m_grade[1][delete-time]
            del m_grade[2][delete-time]
            time += 1
        student_del = []
    return m_grade


def dealwithdata(grade=[[], [], []]):  # 处理数据
    grade = dealwithname(grade)
    grade = checkduplication(grade)
    return grade


if __name__ == "__main__":
    address = sayhello()
    wb = load_workbook(address)
    info = askinfo_start()
    #print(info)
    answer = getanswer()
    ws = wb["Sheet1"]
    point = int(input("请设置分值： "))
    maxcolumn = ws.max_column
    maxrow = ws.max_row
    row = int(info[0])
    column = int(info[1])
    finalgrade = [[], [], []]  # 储存名字和成绩的列表
    print("最大行数为:"+str(maxrow))
    print("最大列数位:"+str(maxcolumn))
    a = (maxcolumn-column)+1
    print("作业中答案列数为："+str(a))
    if(a != len(answer)):
        print("答案位数错误")
        print("失败")
    else:
        while row <= maxrow:
            grade = 0
            column = 7  # 重置列数
            num_ans = 0
            numb_ans_list = 0
            answer_list = []
            while column <= maxcolumn:  # 判断一行的成绩
                a = ws.cell(row, column).value
                answer_list.append(a)
                if(a == answer[num_ans]):
                    grade += point
                else:
                    grade += 0
                num_ans += 1
                column += 1
            ws.cell(row, column).value = grade  # 添加成绩
            finalgrade[0].append(ws.cell(row, 1).value)
            finalgrade[1].append(grade)
            finalgrade[2].append(answer_list)
            print("第"+str(row)+"行成绩为："+str(ws.cell(row, column).value))
            row += 1  # 进1行
        print("最终成绩为: ")
        print(finalgrade)
        print("开始剔除重复")
        fianlgrade = dealwithdata(finalgrade)
        print("剔除结果为")
        print(finalgrade)
        outputexcel(fianlgrade)
        print("完成")

    input()
    pass
