from openpyxl import load_workbook
from openpyxl import Workbook


import os
import json
import types


def check_address(address):  # 检查路径是否标准，文件类型是否支持
    if os.path.isfile(address):
        if address.endswith(".xlsx"):
            print("文件检查正确")
            return True
        else:
            print("文件不支持请重新输入路径")
            return False
    else:
        print("找不到目标文件，请重新输入")
        return False


def set_homework_address():  # 获取作业文件路径
    fileaddress = input("请输入作业路径: ")
    return fileaddress


def set_answer_address():  # 获取答案文件路径
    fileaddress = input("请输入答案路径: ")
    return fileaddress


def get_homework_list(homework_sheet, row_start, column_start, name_column, check_column):  # 将作业转换成列表
    homework_list = [[], [], [], []]  # 这里分别是 0:姓名 1:是否答题 2:成绩 3:答案
    homework_row_max = homework_sheet.max_row
    print("作业最大行数为："+str(homework_row_max))
    homework_column_max = homework_sheet.max_column
    print("作业最大列数为："+str(homework_column_max))
    name_list = get_column_homework(
        homework_sheet, row_start, name_column)  # 获取名字列表
    homework_list[0] = name_list  # 将姓名列表加入作业队列
    answer_list = get_answer_homework(
        homework_sheet, row_start, column_start)  # 获取答案列表
    homework_list[3] = answer_list  # 将答案队列加入作业队列
    check_list = get_column_homework(
        homework_sheet, row_start, check_column)  # 获取是否填写队列
    homework_list[1] = check_list  # 将是否填写队列加入作业队列
    return homework_list


def get_column_homework(homework_sheet, row_start, column):  # 获取一列数据
    if(not column == None):  # 检测是不是没有填写默认检查行
        name_list = []
        row_homework_sheet = row_start  # 初始化行索引
        homework_row_max = homework_sheet.max_row
        while(row_homework_sheet <= homework_row_max):
            name_list.append(homework_sheet.cell(
                row_homework_sheet, column).value)  # 加入列表
            row_homework_sheet += 1  # 指向下一行
        return name_list
    else:
        name_list = []
        row_homework_sheet = row_start  # 初始化行索引
        homework_row_max = homework_sheet.max_row
        while(row_homework_sheet <= homework_row_max):
            name_list.append("否")  # 加入列表
            row_homework_sheet += 1  # 指向下一行
        return name_list


def get_answer_homework(homework_sheet, row_start, column_start):  # 获取学生答案
    answer_list = []
    answer_list_row_index = 0  # 列表行索引
    row_homework_sheet = row_start  # 初始化表格行索引
    homework_row_max = homework_sheet.max_row
    homework_column_max = homework_sheet.max_column
    while(row_homework_sheet <= homework_row_max):
        column_homework_sheet = column_start  # 初始化表格列索引，因为名字占了一行所以
        answer_list_row = []  # 每一行的答案
        while(column_homework_sheet <= homework_column_max):
            value = homework_sheet.cell(
                row_homework_sheet, column_homework_sheet).value  # 获取数据
            value=rectify_vlaue_string(value)  # 纠正数据类型为string
            answer_list_row.append(value)  # 加入行列表
            column_homework_sheet += 1  # 指向下一列
        #print(answer_list_row)
        answer_list.append(answer_list_row)
        row_homework_sheet += 1  # 指向下一行
        answer_list_row_index += 1  # 指向下一行
    return answer_list

def get_anser_list(answer_sheet):  # 获取答案
    answer_list = []
    row_answer_sheet = 1  # 初始化行索引
    answer_row_max = answer_sheet.max_row  # 取得最大长度
    #print("答案表的最大长度为："+str(answer_row_max))
    while(row_answer_sheet <= answer_row_max):
        cell_value = answer_sheet.cell(row_answer_sheet, 1).value
        cell_value = rectify_vlaue_string(cell_value)
        answer_list.append(cell_value)
        row_answer_sheet += 1
    #print("答案列表为：")
    #print(answer_list)
    return answer_list


def get_grade_answer(anwer_sheet):  # 获取每道题的分值
    answer_grade_list = []
    row_answer_sheet = 1  # 初始化行索引
    if(answer_sheet.cell(2, 2).value == None or answer_sheet.cell(2, 2).value == ""):
        answer_row_max = answer_sheet.max_row  # 取得最大长度
        while(row_answer_sheet <= answer_row_max):
            cell_value = answer_sheet.cell(1, 2).value
            cell_value = rectify_vlaue_int(cell_value)
            answer_grade_list.append(cell_value)
            row_answer_sheet += 1
        #print("答案分值列表为：")
        #print(answer_grade_list)
        return answer_grade_list
    else:
        answer_row_max = answer_sheet.max_row  # 取得最大长度
        while(row_answer_sheet <= answer_row_max):
            cell_value = answer_sheet.cell(row_answer_sheet, 2).value
            if(cell_value==None or cell_value==""):
                print("发现答案数分值据缺失，无法继续，请检查答案分值是否完整")
                return None
            cell_value = rectify_vlaue_int(cell_value)
            answer_grade_list.append(cell_value)
            row_answer_sheet += 1
        #print("答案分值列表为：")
        #print(answer_grade_list)
        return answer_grade_list

def rectify_vlaue_string(value):  # 纠正数据类型为string
    if(not type(value) == type("a")):
        value = str(value)
    return value


def rectify_vlaue_int(value):  # 纠正数据类型为int
    if(not type(value) == type(1)):
        value = int(value)
    return value


def load_homework_sheet(address):  # 加载目标文件
    myworkbook = load_workbook(address)
    return myworkbook


def get_work_sheet(workbook):  # 获取默认工作表
    sheet = workbook.active
    return sheet


def delete_target_data(homework_list, target_list):  # 去除特定目标数据
    homework_index = 0
    delete_list = []
    for name in homework_list[0]:
        for target_name in target_list:
            if target_name == name:
                delete_list.append(homework_index)
        homework_index+=1#指向下一个
    for index in delete_list:
        del homework_list[0][index]
        del homework_list[1][index]
        del homework_list[2][index]
        del homework_list[3][index]
    return homework_list

def split_name(homework_list,split_list):#精简名字
    name_index = 0#初始化名字索引
    for name in homework_list[0]:
        name.rstrip()
        for a in split_list:
            if name.find(a) >= 0:
                newname = name.split(a)
                homework_list[0][name_index] = newname[0]
                break
        name_index += 1
    return homework_list

def delete_duplication_data(homework_list):  # 处理重复的数据
    index_row = 0#初始化行索引
    while(index_row<=len(homework_list[0])):
        offset=1#初始化偏移量
        delete_list = []#初始化要删除的位置列表
        #length=len(homework_list[0])//debug用
        while((index_row+offset+1)<=len(homework_list[0])):
            if(homework_list[0][index_row]==homework_list[0][index_row+offset]):#判断是否相等
                if(homework_list[1][index_row+offset]=="否"):#判断是否答题
                    delete_list.append((index_row+offset))#将索引值加入删除名单
            offset+=1#指向下一个
        delete_offset=0#纠正删除的偏移量
        for index in delete_list:#删除应该删除的数据
            del homework_list[0][index-delete_offset]
            del homework_list[1][index-delete_offset]
            del homework_list[2][index-delete_offset]
            del homework_list[3][index-delete_offset]
            delete_offset+=1
        index_row+=1#指向下一行
    return homework_list


def compute_grade(homework_list, answer_list, grade_list):  # 计算成绩
    if(grade_list==None):
        return None
    for row in homework_list[3]:
        grade_row = compute_grade_row(row, answer_list, grade_list)  # 计算单行成绩
        homework_list[2].append(grade_row)  # 写入单行成绩
    return homework_list


def compute_grade_row(homework_row_list, answer_list, grade_list):  # 计算单行成绩
    answer_index = 0
    grade_row = 0
    for answer in homework_row_list:
        if(answer == answer_list[answer_index]):
            grade_row += grade_list[answer_index]
        answer_index += 1  # 指向下一个
    return grade_row


def load_json(address):  # 加载json默认设置文件
    with open("config.json") as json_file:
        config_dict = json.load(json_file)
    return config_dict


def create_json():  # 生成json文件
    config_dict = {"默认起始行": 2,
                   "默认起始列": 7,
                   "默认姓名列": 1,
                   "默认是否答题列":6,
                   "剔除名单": [],
                   "切除关键词": []}
    with open("config.json", "w") as json_file:
        json.dump(config_dict, json_file,ensure_ascii=False)

    json_file.close()


def is_first_setup():  # 判断是否第一次启动
    if not os.path.isfile("config.json"):
        return True
    else:
        return False


def applicate_setting():  # 配置设置
    target_list = config_dict["剔除名单"]
    split_list = config_dict["切除关键词"]
    row_start = config_dict["默认起始行"]
    column_start = config_dict["默认起始列"]
    name_column = config_dict["默认姓名列"]
    check_column = config_dict["默认是否答题列"]


def check_answer_row(homework_sheet, column_start, answer_list):  # 检查答案与作业是否匹配
    print("答案长度: "+str(len(answer_list)))
    print("作业长度: "+str((homework_sheet.max_column-column_start)+1))
    if(((homework_sheet.max_column-column_start)+1) == len(answer_list)):
        return True
    else:
        return False

def output_excel(homework_list):#输出表格
    row = 1#初始化行
    newwb = Workbook()
    newws = newwb.active
    for name in homework_list[0]:#输出名字
        newws.cell(row, 1).value = name
        row += 1
    row = 1#初始化行
    for grade_final in homework_list[2]:#输出成绩
        newws.cell(row, 2).value = grade_final
        row += 1
    row = 1#初始化行
    for answer in homework_list[3]:#输出答案
        column = 3#初始化列
        for answer_s in answer:#横向输出
            newws.cell(row, column).value = answer_s
            column += 1
        row += 1
    address = input("请输入输出地址：")
    if(not address.endswith(".xlsx")):#纠正输入错误
        address=address+".xlsx"
    newwb.save(address)

if __name__ == "__main__":
    print("欢迎使用")

    split_list = []  # 剔除人的名单，全局变量
    row_start = None
    column_start = None
    name_column = None

    if is_first_setup():
        print("检测到您是第一次使用，正在生成相应配置文件")
        create_json()
    #配置全局变量设置
    config_dict = load_json("config.json")
    target_list = config_dict["剔除名单"]
    print("剔除名单:")
    print(target_list)
    split_list = config_dict["切除关键词"]
    print("切除关键词：")
    print(split_list)
    row_start = config_dict["默认起始行"]
    print("默认起始行："+str(row_start))
    column_start = config_dict["默认起始列"]
    print("默认起始列："+str(column_start))
    name_column = config_dict["默认姓名列"]
    print("默认姓名列："+str(name_column))
    check_column = config_dict["默认是否答题列"]
    print("默认是否答题列："+str(check_column))

    #输入作业地址并检测
    while(True):
        homework_address = set_homework_address()
        if check_address(homework_address):
            break
    homework_workbook = load_workbook(homework_address)  # 获取作业表格
    homework_sheet = get_work_sheet(homework_workbook)  # 获取作业工作表
    homework_list = get_homework_list(
        homework_sheet, row_start, column_start, name_column, check_column)  # 获取作业列表

    #输入答案地址并检测
    while(True):
        answer_address = set_answer_address()
        if check_address(answer_address):
            break
    answer_workbook = load_workbook(answer_address)  # 获取答案表格
    answer_sheet = get_work_sheet(answer_workbook)  # 获取答案工作表
    answer_list = get_anser_list(answer_sheet)  # 获取答案列表
    grade_list = get_grade_answer(answer_sheet)  # 获取每道题的分值

    #循环验证答案与作业是否匹配
    while(True):
        if(not check_answer_row(homework_sheet, column_start, answer_list)):
            print("答案位数错误")
            #输入答案地址并检测
            while(True):
                answer_address = set_answer_address()
                if check_address(answer_address):
                    break
            answer_workbook = load_workbook(answer_address)  # 获取答案表格
            answer_sheet = get_work_sheet(answer_workbook)  # 获取答案工作表
            answer_list = get_anser_list(answer_sheet)  # 获取答案列表
            grade_list = get_grade_answer(answer_sheet)  # 获取每道题的分值
        else:
            print("答案检验对照正确")
            break

    #开始计算成绩
    print("开始计算成绩")
    homework_list = compute_grade(homework_list, answer_list, grade_list)
    if(homework_list==None):#判断是否有错误发生
        print("程序结束")
        input()
    else:
        #去除特定人
        print("去除特定人")
        delete_target_data(homework_list, target_list)
        #精简名字
        print("精简名字")
        split_name(homework_list,split_list)
        #删除重复名字
        print("删除重复名字")
        delete_duplication_data(homework_list)
        #输出表格
        print("输出表格")
        output_excel(homework_list)

        print("完成")
        input()
    pass
