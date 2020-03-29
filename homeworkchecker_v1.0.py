from openpyxl import load_workbook
import os

#需要设置分名字系统
#单独设置分值
#

def sayhello():
    print("程序开始执行")
    filename = input("请输入文件路径: ")
    address = filename
    #print("检测到文件路径为: "+address)
    return address


def askinfo_start():
    info_start = []
    row_start = input("请输入起始行： ")
    if row_start == None:
        row_start = 1
    info_start.append(row_start)
    column_start = input("请输入起始列：")
    if column_start == None:
        column_start = 1
    info_start.append(column_start)
    return info_start


def getanswer():
    answeraddress = input("请输入答案路径： ")
    #print("检测到答案地址为： "+answeraddress)
    wb = load_workbook(answeraddress)
    ws = wb["Sheet1"]
    answer = []
    column = 1
    checkpoint = True
    while checkpoint:
        a = ws.cell(column, 1).value
        if(a == None):
            break
        else:
            answer.append(a)
        column = column+1
    print("答案是：")
    print(answer)
    return answer


if __name__ == "__main__":
    address = sayhello()
    wb = load_workbook(address)
    info = askinfo_start()
    #print(info)
    answer = getanswer()
    ws = wb["Sheet1"]
    point = int(input("请设置分值： "))
    maxcolumn = ws.max_column
    maxrow = ws.max_row
    row = int(info[0])
    column = int(info[1])
    print("最大行数为:"+str(maxrow))
    print("最大列数位:"+str(maxcolumn))
    a = (maxcolumn-column)+1
    print("作业中答案列数为："+str(a))
    if(a != len(answer)):
        print("答案位数错误")
    else:
        while row <= maxrow:
            grade = 0
            column = 7  # 重置列数
            num_ans = 0
            while column <= maxcolumn:  # 判断一行的成绩
                a = ws.cell(row, column).value
                if(a == answer[num_ans]):
                    grade += point
                else:
                    grade += 0
                num_ans += 1
                column += 1
            ws.cell(row, column).value = grade  # 添加成绩
            print("第"+str(row)+"行成绩为："+str(ws.cell(row, column).value))

            row += 1  # 进1行
        saveaddr = input("请输入储存地址： ")
        wb.save(saveaddr)
        print("完成")
    print("失败")
    input()
    pass
